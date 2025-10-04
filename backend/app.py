import hashlib
import io
import json
import os
import re
import smtplib
import ssl
import time
import threading
from datetime import datetime, timedelta
from email.message import EmailMessage
from urllib.parse import urljoin, urlparse
from functools import wraps

import requests
from bs4 import BeautifulSoup
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright
import jwt
from werkzeug.security import check_password_hash, generate_password_hash

app = Flask(__name__)
# Define the specific origin of your frontend app
frontend_origin = "https://pci-app.theironshield.ca" 

CORS(app, resources={
    r"/api/*": {
        "origins": [frontend_origin, "http://localhost:3000"], # Allow both production and local dev
        "methods": ["GET", "POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "supports_credentials": True
    }
})

# --------------------- Authentication Configuration ---------------------
SECRET_KEY = os.environ.get("SECRET_KEY", "your-secret-key-change-in-production")
TOKEN_EXPIRATION_HOURS = 24

USERS_DB = {
    "admin": generate_password_hash("admin"),
    "user": generate_password_hash("password123"),
}

# --------------------- Cross-platform timestamp formatting ---------------------
def format_ts(dt: datetime) -> str:
    date_part = f"{dt.month}/{dt.day}/{dt.year}"
    time_part = dt.strftime("%I:%M:%S %p").lstrip("0")
    return f"{date_part}, {time_part}"

# --------------------- Email Configuration ---------------------
SMTP_HOST = os.environ.get("SMTP_HOST")
SMTP_PORT = int(os.environ.get("SMTP_PORT") or "587")
SMTP_USER = os.environ.get("SMTP_USER")
SMTP_PASS = os.environ.get("SMTP_PASS")
SMTP_FROM = os.environ.get("SMTP_FROM")

def send_alert_email(subject: str, body: str):
    recipients = MOCK_DB["page_config"].get("email_recipients", []) or []
    if not (SMTP_HOST and SMTP_FROM and recipients):
        return
    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = SMTP_FROM
        msg["To"] = ", ".join(recipients)
        msg.set_content(body)
        context = ssl.create_default_context()
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls(context=context)
            if SMTP_USER and SMTP_PASS:
                server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)
    except Exception as e:
        print(f"[EMAIL] Failed to send alert: {e}")

def send_slack_alert(message: str):
    slack_url = MOCK_DB["notifications"]["slack_url"]
    if not slack_url:
        return
    try:
        payload = {"text": message}
        requests.post(slack_url, json=payload, timeout=10)
    except Exception as e:
        print(f"[SLACK] Failed to send alert: {e}")

def send_teams_alert(message: str):
    teams_url = MOCK_DB["notifications"]["teams_url"]
    if not teams_url:
        return
    try:
        payload = {"text": message}
        requests.post(teams_url, json=payload, timeout=10)
    except Exception as e:
        print(f"[TEAMS] Failed to send alert: {e}")

def broadcast_alert(subject: str, body: str):
    send_alert_email(subject, body)
    full_message = f"*{subject}*\n\n{body}"
    send_slack_alert(full_message)
    send_teams_alert(full_message)

# --------------------- In-memory state ---------------------
MOCK_DB = {
    "page_config": {
        "url": "https://example.com/",
        "email_recipients": [],
        "compliance_status": "N/A",
        "last_scan": "N/A",
        "scope": "all",
    },
    "notifications": {
        "scan_interval": 3600,
        "slack_url": "",
        "teams_url": "",
    },
    "scripts": {},
    "alerts": [],
    "alert_seen": {},
    "baseline": {},
    "tamper_cache": {},
}
DEDUP_WINDOW_SEC = 30

TRUSTED_VOLATILE_HOSTS = {
    "pay.google.com",
    "www.googletagmanager.com",
    "static.checkout.littlepay.com",
    "verify.littlepay.com",
}
AUTO_REBASELINE_TRUSTED = True

# --------------------- Authentication Decorator ---------------------
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(" ")[1]
            except IndexError:
                return jsonify({"message": "Invalid token format"}), 401
        
        if not token:
            return jsonify({"message": "Token is missing"}), 401
        
        try:
            data = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            current_user = data['username']
        except jwt.ExpiredSignatureError:
            return jsonify({"message": "Token has expired"}), 401
        except jwt.InvalidTokenError:
            return jsonify({"message": "Invalid token"}), 401
        
        return f(current_user, *args, **kwargs)
    
    return decorated

# --------------------- Helpers ---------------------
def sha256_hex(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()

def is_external_src(src: str) -> bool:
    return bool(re.match(r"^https?://", src or ""))

def classify_script(page_url: str, src: str | None, is_inline: bool) -> str:
    if is_inline:
        return "Inline"
    if not src:
        return "Internal"
    page_host = urlparse(page_url).netloc or ""
    src_host = urlparse(src).netloc or ""
    return "Vendor" if src_host and src_host != page_host else "Internal"

def fetch_script_bytes(resolved_url: str) -> bytes:
    try:
        resp = requests.get(resolved_url, timeout=15)
        resp.raise_for_status()
        return resp.content
    except Exception:
        return b""

def check_sri_on_tag(tag) -> bool:
    if not tag:
        return True
    src = tag.get("src")
    if not src:
        return True
    return bool(tag.get("integrity"))

def filename_from_url(url: str, is_inline: bool = False) -> str:
    if is_inline or not url:
        return "inline"
    try:
        p = urlparse(url)
        name = (p.path or "").rstrip("/").split("/")[-1] or p.netloc or url
        return name
    except Exception:
        return url

# --------------------- Browser render ---------------------
def get_page_rendered(url: str, settle_ms: int = 3000, include_subframes: bool = True):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        js_requests = set()

        def on_request_finished(req):
            if not include_subframes and req.frame and req.frame != page.main_frame:
                return
            try:
                res = req.response()
                ct = (res.headers or {}).get("content-type", "")
            except Exception:
                ct = ""
            url_no_qs = req.url.split("?")[0]
            looks_js = (
                req.resource_type in ("script", "xhr", "fetch")
                and ("javascript" in ct or url_no_qs.endswith((".js", ".mjs")))
            )
            if looks_js:
                js_requests.add(req.url)

        page.on("requestfinished", on_request_finished)

        page.goto(url, wait_until="domcontentloaded")
        time.sleep(settle_ms / 1000.0)
        try:
            page.wait_for_load_state("networkidle", timeout=5000)
        except Exception:
            pass

        final_url = page.url
        html = page.content()

        dom_script_urls = set()
        frames = page.frames if include_subframes else [page.main_frame]
        for frame in frames:
            for handle in frame.locator("script").element_handles():
                try:
                    src = handle.get_attribute("src")
                except Exception:
                    src = None
                if src:
                    dom_script_urls.add(urljoin(frame.url, src))

        browser.close()

    resp = requests.get(final_url, timeout=15)
    resp.raise_for_status()
    headers = resp.headers

    all_js_urls = sorted(dom_script_urls.union(js_requests))
    return html, headers, final_url, all_js_urls

# --------------------- Alerts / tamper ---------------------
def raise_alert(alert: dict):
    key = (alert.get("type"), alert.get("location"), alert.get("severity"), alert.get("details"))
    now = time.time()

    last = MOCK_DB["alert_seen"].get(key)
    if last and (now - last) < DEDUP_WINDOW_SEC:
        return
    MOCK_DB["alert_seen"][key] = now

    for a in MOCK_DB["alerts"]:
        if (
            a.get("type") == alert.get("type") and
            a.get("location") == alert.get("location") and
            a.get("severity") == alert.get("severity") and
            a.get("details") == alert.get("details")
        ):
            a["timestamp"] = alert["timestamp"]
            break
    else:
        MOCK_DB["alerts"].insert(0, alert)
        if alert.get("severity") == "HIGH":
            subject = f"[PCI Monitor] {alert.get('type')} at {urlparse(alert.get('location','')).netloc or 'page'}"
            body = (
                f"Type: {alert.get('type')}\n"
                f"Severity: {alert.get('severity')}\n"
                f"Time: {alert.get('timestamp')}\n"
                f"Location: {alert.get('location')}\n"
                f"Details: {alert.get('details')}\n"
            )
            broadcast_alert(subject, body)

def record_tamper_if_changed(location: str, new_hash: str):
    baseline = MOCK_DB["baseline"].get(location)
    if baseline is None:
        MOCK_DB["baseline"][location] = new_hash
        return
    if baseline != new_hash:
        last_alerted = MOCK_DB["tamper_cache"].get(location)
        if last_alerted == new_hash:
            return
        MOCK_DB["tamper_cache"][location] = new_hash
        host = urlparse(location).netloc
        sev = "MEDIUM" if host in TRUSTED_VOLATILE_HOSTS else "HIGH"
        raise_alert({
            "id": len(MOCK_DB["alerts"]) + 1,
            "timestamp": datetime.now().isoformat(),
            "type": "Script Tampered",
            "location": location,
            "details": f"Hash changed from {baseline[:12]}… to {new_hash[:12]}….",
            "severity": sev
        })
        if AUTO_REBASELINE_TRUSTED and host in TRUSTED_VOLATILE_HOSTS:
            MOCK_DB["baseline"][location] = new_hash

# --------------------- Scanner ---------------------
def run_scan(url: str, include_subframes: bool = True):
    html, headers, final_url, all_js_urls = get_page_rendered(
        url, include_subframes=include_subframes
    )

    csp_value = headers.get("Content-Security-Policy")
    MOCK_DB["scripts"]["headers"] = {
        "id": "headers",
        "type": "Headers",
        "location": final_url,
        "sri_applied": True,
        "hash": sha256_hex(json.dumps(dict(headers), sort_keys=True).encode()),
        "content": dict(headers),
    }
    if not csp_value:
        raise_alert({
            "id": len(MOCK_DB["alerts"]) + 1,
            "timestamp": datetime.now().isoformat(),
            "type": "CSP Missing",
            "location": final_url,
            "details": "No Content-Security-Policy header present.",
            "severity": "MEDIUM",
        })

    soup = BeautifulSoup(html, "html.parser")
    dom_seen = set()
    dom_script_tags = soup.find_all("script")

    for idx, tag in enumerate(dom_script_tags, start=1):
        src = tag.get("src")
        is_inline = src is None
        resolved = urljoin(final_url, src) if src else f"{final_url}#inline-{idx}"
        dom_seen.add(resolved)

        sri_ok = check_sri_on_tag(tag)
        content_bytes = (
            fetch_script_bytes(resolved)
            if src
            else (tag.string or tag.text or "").encode("utf-8", errors="ignore")
        )
        script_hash = sha256_hex(content_bytes)

        item_id = f"script-{idx}"
        MOCK_DB["scripts"][item_id] = {
            "id": item_id,
            "type": classify_script(final_url, resolved if src else None, is_inline),
            "location": resolved,
            "sri_applied": bool(sri_ok),
            "hash": script_hash,
        }

        if src and is_external_src(resolved) and not sri_ok:
            raise_alert({
                "id": len(MOCK_DB["alerts"]) + 1,
                "timestamp": datetime.now().isoformat(),
                "type": "SRI Missing",
                "location": resolved,
                "details": "Script loaded without Subresource Integrity.",
                "severity": "MEDIUM",
            })

        record_tamper_if_changed(resolved, script_hash)

    dynamic_only = [u for u in all_js_urls if u not in dom_seen]
    for j, js_url in enumerate(dynamic_only, start=1):
        content_bytes = fetch_script_bytes(js_url)
        script_hash = sha256_hex(content_bytes)
        item_id = f"dynamic-{j}"
        MOCK_DB["scripts"][item_id] = {
            "id": item_id,
            "type": classify_script(final_url, js_url, False),
            "location": js_url,
            "sri_applied": False,
            "hash": script_hash,
        }
        if is_external_src(js_url):
            raise_alert({
                "id": len(MOCK_DB["alerts"]) + 1,
                "timestamp": datetime.now().isoformat(),
                "type": "SRI Missing",
                "location": js_url,
                "details": "Dynamically loaded script without Subresource Integrity.",
                "severity": "MEDIUM",
            })
        record_tamper_if_changed(js_url, script_hash)

    overall_alert = any(a.get("severity") in ("HIGH", "MEDIUM") for a in MOCK_DB["alerts"])
    MOCK_DB["page_config"]["compliance_status"] = "ALERT" if overall_alert else "OK"
    MOCK_DB["page_config"]["last_scan"] = format_ts(datetime.now())
    MOCK_DB["page_config"]["url"] = final_url

    return {
        "status": MOCK_DB["page_config"]["compliance_status"],
        "last_scan": MOCK_DB["page_config"]["last_scan"],
        "count_scripts": len(dom_script_tags) + len(dynamic_only),
        "has_csp": bool(csp_value),
    }

# --------------------- Periodic Scanner ---------------------
def send_summary_report():
    status = MOCK_DB["page_config"]["compliance_status"]
    last_scan = MOCK_DB["page_config"]["last_scan"]
    url = MOCK_DB["page_config"]["url"]
    unresolved_alerts = len(MOCK_DB["alerts"])
    subject = f"PCI Scan Report for {urlparse(url).netloc}: {status}"
    body = (
        f"Scan completed at: {last_scan}\n"
        f"URL Scanned: {url}\n"
        f"Overall Compliance Status: {status}\n"
        f"Total Unresolved Alerts: {unresolved_alerts}\n\n"
    )
    if unresolved_alerts > 0:
        body += "Recent Alerts:\n"
        for alert in MOCK_DB["alerts"][:5]:
            body += f"- {alert['type']} ({alert['severity']}) at {alert['location']}\n"
    broadcast_alert(subject, body)
    print("[REPORTER] Summary report sent to all channels.")

def periodic_scanner():
    print("[SCANNER] Starting periodic scanner thread.")
    while True:
        interval = MOCK_DB["notifications"]["scan_interval"]
        try:
            url_to_scan = MOCK_DB["page_config"]["url"]
            if url_to_scan and interval > 0:
                print(f"[SCANNER] Starting scheduled scan for {url_to_scan}...")
                with app.app_context():
                    include_subframes = (MOCK_DB["page_config"]["scope"] == "all")
                    run_scan(url_to_scan, include_subframes=include_subframes)
                    send_summary_report()
                print("[SCANNER] Scheduled scan finished.")
        except Exception as e:
            print(f"[SCANNER] Error during scheduled scan: {e}")
        
        sleep_duration = interval if interval > 0 else 86400
        print(f"[SCANNER] Next scan in {sleep_duration} seconds.")
        time.sleep(sleep_duration)

# --------------------- Authentication API ---------------------
@app.route("/api/login", methods=["POST"])  # NO @token_required here!
def login():
    data = request.get_json()
    username = data.get("username")
    password = data.get("password")
    
    print(f"[LOGIN] Received - Username: '{username}', Password: '{password}'")  # Debug log
    
    if not username or not password:
        return jsonify({"success": False, "message": "Username and password required"}), 400
    
    if username in USERS_DB and check_password_hash(USERS_DB[username], password):
        token = jwt.encode({
            'username': username,
            'exp': datetime.utcnow() + timedelta(hours=TOKEN_EXPIRATION_HOURS)
        }, SECRET_KEY, algorithm="HS256")
        
        return jsonify({
            "success": True,
            "token": token,
            "username": username,
            "message": "Login successful"
        })
    
    print(f"[LOGIN] Failed - Invalid credentials for user: {username}")  # Debug log
    return jsonify({"success": False, "message": "Invalid credentials"}), 401

@app.route("/api/validate-token", methods=["POST"])
@token_required
def validate_token(current_user):
    return jsonify({
        "valid": True,
        "username": current_user
    })

# --------------------- Protected API Endpoints ---------------------
@app.route("/api/status", methods=["GET"])
@token_required
def get_status(current_user):
    return jsonify({
        "overall": MOCK_DB["page_config"]["compliance_status"],
        "last_scan": MOCK_DB["page_config"]["last_scan"],
        "unresolved": len(MOCK_DB["alerts"]),
        "scope": MOCK_DB["page_config"].get("scope", "all"),
    })

@app.route("/api/inventory", methods=["GET"])
@token_required
def get_inventory(current_user):
    return jsonify(list(MOCK_DB["scripts"].values()))

@app.route("/api/alerts", methods=["GET"])
@token_required
def get_alerts(current_user):
    return jsonify(MOCK_DB["alerts"])

def _resolve_scope_from_payload(payload: dict) -> str:
    scope = (payload or {}).get("scope")
    if scope in ("all", "entry"):
        return scope
    return MOCK_DB["page_config"].get("scope", "all")

@app.route("/api/scan", methods=["POST"])
@token_required
def manual_scan(current_user):
    data = request.get_json(silent=True) or {}
    url = data.get("url") or MOCK_DB["page_config"]["url"]
    scope = _resolve_scope_from_payload(data)
    MOCK_DB["page_config"]["scope"] = scope
    include_subframes = (scope == "all")
    try:
        result = run_scan(url, include_subframes=include_subframes)
        return jsonify({"message": "Scan complete", "result": result, "scope": scope})
    except Exception as e:
        raise_alert({
            "id": len(MOCK_DB["alerts"]) + 1,
            "timestamp": datetime.now().isoformat(),
            "type": "Scan Failure",
            "location": url,
            "details": str(e),
            "severity": "MEDIUM",
        })
        MOCK_DB["page_config"]["compliance_status"] = "ALERT"
        MOCK_DB["page_config"]["last_scan"] = format_ts(datetime.now())
        return jsonify({"message": "Scan failed", "error": str(e)}), 500

@app.route("/api/reset", methods=["POST"])
@token_required
def reset_and_rescan(current_user):
    data = request.get_json(silent=True) or {}
    url = data.get("url") or MOCK_DB["page_config"]["url"]
    scope = _resolve_scope_from_payload(data)
    MOCK_DB["page_config"]["scope"] = scope
    include_subframes = (scope == "all")

    MOCK_DB["alerts"].clear()
    MOCK_DB["scripts"].clear()
    MOCK_DB["baseline"].clear()
    MOCK_DB["tamper_cache"].clear()
    MOCK_DB["page_config"]["compliance_status"] = "N/A"
    MOCK_DB["page_config"]["last_scan"] = "N/A"
    try:
        result = run_scan(url, include_subframes=include_subframes)
        return jsonify({"message": "State cleared and fresh scan complete", "result": result, "scope": scope})
    except Exception as e:
        return jsonify({"message": "Reset complete but scan failed", "error": str(e)}), 500

@app.route("/api/hard-reset", methods=["POST"])
@token_required
def hard_reset(current_user):
    MOCK_DB["alerts"].clear()
    MOCK_DB["scripts"].clear()
    MOCK_DB["baseline"].clear()
    MOCK_DB["tamper_cache"].clear()
    MOCK_DB["page_config"] = {
        "url": "",
        "email_recipients": [],
        "compliance_status": "N/A",
        "last_scan": "N/A",
        "scope": "all"
    }
    return jsonify({"message": "Application state has been cleared."})

def _parse_emails(value):
    if isinstance(value, list):
        parts = value
    else:
        parts = [p.strip() for p in (value or "").split(",")]
    return [p for p in parts if p and "@" in p]

@app.route("/api/config", methods=["GET"])
@token_required
def get_config(current_user):
    cfg = MOCK_DB["page_config"]
    return jsonify({
        "url": cfg.get("url"),
        "email_recipients": cfg.get("email_recipients", []),
        "scope": cfg.get("scope", "all"),
    })

@app.route("/api/config", methods=["POST"])
@token_required
def save_config(current_user):
    data = request.get_json() or {}
    url = data.get("url")
    emails = data.get("emails", [])
    scope = data.get("scope")

    if not url or not re.match(r"^https?://", url):
        return jsonify({"message": "Invalid URL"}), 400

    MOCK_DB["page_config"]["url"] = url
    MOCK_DB["page_config"]["email_recipients"] = _parse_emails(emails)
    if scope in ("all", "entry"):
        MOCK_DB["page_config"]["scope"] = scope

    include_subframes = (MOCK_DB["page_config"]["scope"] == "all")
    try:
        result = run_scan(url, include_subframes=include_subframes)
        return jsonify({"message": "Configuration saved", "result": result, "scope": MOCK_DB['page_config']['scope']})
    except Exception as e:
        return jsonify({"message": "Configuration saved, but scan failed", "error": str(e)}), 200

@app.route("/api/notifications-config", methods=["GET"])
@token_required
def get_notifications_config(current_user):
    return jsonify(MOCK_DB["notifications"])

@app.route("/api/notifications-config", methods=["POST"])
@token_required
def save_notifications_config(current_user):
    data = request.get_json() or {}
    try:
        interval = int(data.get("scan_interval", 3600))
        MOCK_DB["notifications"]["scan_interval"] = interval
    except (ValueError, TypeError):
        return jsonify({"message": "Invalid scan_interval, must be a number."}), 400
    
    MOCK_DB["notifications"]["slack_url"] = data.get("slack_url", "")
    MOCK_DB["notifications"]["teams_url"] = data.get("teams_url", "")
    
    return jsonify({"message": "Notification settings saved."})

@app.route("/api/export/inventory.xlsx", methods=["GET"])
@token_required
def export_inventory_excel(current_user):
    rows = []
    for item in MOCK_DB["scripts"].values():
        if item.get("id") == "headers":
            continue
        loc = item.get("location", "")
        raw_type = item.get("type", "")
        type_label = {"Internal": "First-party", "Vendor": "Third-party", "Inline": "Inline"}.get(raw_type, raw_type)
        is_inline = (raw_type == "Inline")
        rows.append([
            filename_from_url(loc, is_inline),
            loc,
            type_label,
            "Applied" if item.get("sri_applied") else "Missing",
            item.get("hash", ""),
        ])
    if not rows:
        rows = [["(none)", "", "", "", ""]]

    wb = Workbook()
    ws = wb.active
    ws.title = "Scripts"
    headers = ["Script", "URL", "Type", "SRI", "Hash (sha256)"]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    widths = [30, 80, 14, 10, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"script_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# --------------------- Helper Functions ---------------------
def add_user(username, password):
    """Add a new user to the system"""
    USERS_DB[username] = generate_password_hash(password)
    print(f"User '{username}' added successfully")

# --------------------- Application Startup ---------------------
if __name__ == "__main__":
    if os.environ.get("SCAN_ON_START") == "1":
        try:
            with app.app_context():
                run_scan(MOCK_DB["page_config"]["url"], include_subframes=True)
        except Exception:
            pass
    
    scan_thread = threading.Thread(target=periodic_scanner, daemon=True)
    scan_thread.start()
    
    app.run(host="0.0.0.0", port=5000, debug=True)